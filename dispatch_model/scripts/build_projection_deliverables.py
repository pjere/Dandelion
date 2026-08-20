"""Turn the persisted 20-year projection into the four requested deliverables.

Reads the parquet set written by `run_projection_20y.py` and emits, into `--out`:

    projection_20y_spot_prices.xlsx     hourly spot per zone over the whole horizon (+ FR SMC)
    projection_20y_analysis.xlsx        yearly revenue by tech/zone, FR congestion rent, negative hours
    *.csv                               the same tables, for anything that would rather not open Excel

Two price series run through every table, and the distinction matters:

  * **SMC** is the LP's own balance dual — the model's internally consistent price. Congestion rent is
    non-negative by construction on SMC, because flow only ever runs from the cheaper zone to the dearer.
  * **spot** is SMC after the step-vii markup, clipped to [-500, 4000]. It is the model's estimate of what
    a market would actually print, and it is what a generator is paid — so revenues lead on spot. But the
    markup is a per-zone regression extrapolated 20 years past its fit, and the clip at 4000 bites whenever
    the LP reaches its 15000 VOLL, so both series are reported side by side rather than one being hidden.

Volumes come from the LP primal (`DISPATCH_CAPTURE_DISPATCH`), so revenue is generation x price hour by
hour, not an annual average times an annual volume — the difference is the capture-rate effect, which for
solar and wind is the entire economic question.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

import numpy as np
import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

# pseudo-columns that are not generation technologies
_SYSTEM = ("ens", "dump")
_STORAGE_SINK = "storage_charge"


def _load_year(src: Path, y: int):
    px = pd.read_parquet(src / f"prices_{y}.parquet")
    smc, spot = px["smc"], px["spot"]
    d = None
    if (src / f"dispatch_{y}.parquet").exists():
        d = pd.read_parquet(src / f"dispatch_{y}.parquet")
        d.columns = pd.MultiIndex.from_tuples([tuple(c.split("|", 1)) for c in d.columns],
                                              names=["zone", "tech"])
    fl = None
    if (src / f"flows_{y}.parquet").exists():
        fl = pd.read_parquet(src / f"flows_{y}.parquet")
        fl["time"] = pd.to_datetime(fl["time"], utc=True)
        fl = fl.pivot_table(index="time", columns="border", values="flow_mw", aggfunc="sum") \
               .reindex(smc.index)
    return smc, spot, d, fl


def revenues(y, smc, spot, d) -> pd.DataFrame:
    """Yearly generation and revenue per (zone, tech). Revenue = sum_h MW_h x price_h (MWh x EUR/MWh)."""
    rows = []
    for (z, tech) in d.columns:
        if tech in _SYSTEM:
            continue
        mw = d[(z, tech)]
        vol = float(mw.sum())                                   # MWh (hourly steps)
        sign = -1.0 if tech == _STORAGE_SINK else 1.0           # charging is a purchase, not a sale
        r_spot = float((mw * spot[z]).sum()) * sign
        r_smc = float((mw * smc[z]).sum()) * sign
        if vol <= 0 and abs(r_spot) < 1.0:
            continue
        rows.append({"year": y, "zone": z, "tech": tech,
                     "generation_TWh": vol / 1e6,
                     "revenue_MEUR_spot": r_spot / 1e6,
                     "capture_price_spot": (r_spot / vol) if vol > 0 else np.nan,
                     "revenue_MEUR_smc": r_smc / 1e6,
                     "capture_price_smc": (r_smc / vol) if vol > 0 else np.nan,
                     "zone_mean_spot": float(spot[z].mean()),
                     "capture_rate_spot": ((r_spot / vol) / float(spot[z].mean())
                                           if vol > 0 and spot[z].mean() != 0 else np.nan)})
    return pd.DataFrame(rows)


def congestion(y, smc, spot, fl, hub="FR") -> pd.DataFrame:
    """Congestion rent per interconnector: sum_h flow_h x (price_importing - price_exporting).

    On SMC this is the LP's own scarcity rent on the NTC row and cannot be negative. On spot it can be,
    because a per-zone markup can reorder two zones the LP had ordered the other way; that count is
    reported rather than hidden, as it measures how far the markup moves the coupling."""
    rows = []
    for b in fl.columns:
        a, z = b.split(">")
        if hub not in (a, z):
            continue
        f = fl[b]
        hourly_smc = f * (smc[z] - smc[a])
        hourly_spot = f * (spot[z] - spot[a])
        # A border is congested exactly when its NTC row binds, and the observable signature of that is a
        # non-zero price spread: uncongested, the two zones converge to one marginal (up to the epsilon
        # loop-flow penalty). Testing the FLOW against its annual maximum instead would be wrong — the cap
        # is hourly and varies by a factor of three across the year, so a fully-loaded hour on a low-NTC
        # hour reads as slack. (That error is what first made Portugal's scarcity look artefactual.)
        rows.append({"year": y, "border": f"{hub}-{z if a == hub else a}",
                     "direction_convention": f"+ = {a} to {z}",
                     "net_flow_TWh": float(f.sum()) / 1e6,
                     "exports_TWh": float(f.clip(lower=0).sum()) / 1e6,
                     "imports_TWh": float(-f.clip(upper=0).sum()) / 1e6,
                     "congestion_rent_MEUR_smc": float(hourly_smc.sum()) / 1e6,
                     "congestion_rent_MEUR_spot": float(hourly_spot.sum()) / 1e6,
                     "hours_congested": int(((smc[z] - smc[a]).abs() > 0.01).sum()),
                     "hours_negative_rent_spot": int((hourly_spot < -1e-6).sum())})
    return pd.DataFrame(rows)


def negative_hours(y, smc, spot) -> pd.DataFrame:
    rows = []
    for z in spot.columns:
        p, q = spot[z], smc[z]
        rows.append({"year": y, "zone": z, "hours": int(len(p)),
                     "hours_le_zero_spot": int((p <= 0).sum()),
                     "hours_lt_zero_spot": int((p < 0).sum()),
                     "hours_eq_zero_spot": int((p == 0).sum()),
                     "mean_negative_price_spot": float(p[p < 0].mean()) if (p < 0).any() else np.nan,
                     "min_price_spot": float(p.min()),
                     "hours_le_zero_smc": int((q <= 0).sum()),
                     "hours_lt_zero_smc": int((q < 0).sum())})
    return pd.DataFrame(rows)


def price_summary(y, smc, spot) -> pd.DataFrame:
    rows = []
    for z in spot.columns:
        p, q = spot[z], smc[z]
        rows.append({"year": y, "zone": z,
                     "mean_spot": p.mean(), "median_spot": p.median(),
                     "p10_spot": p.quantile(.10), "p90_spot": p.quantile(.90),
                     "max_spot": p.max(), "min_spot": p.min(),
                     "mean_smc": q.mean(), "median_smc": q.median(), "max_smc": q.max(),
                     "hours_gt_500_spot": int((p > 500).sum()),
                     "hours_at_clip_4000_spot": int((p >= 3999.99).sum())})
    return pd.DataFrame(rows)


def system_table(y, d) -> pd.DataFrame:
    rows = []
    for z in sorted(set(d.columns.get_level_values("zone"))):
        sub = d[z]
        ens = sub["ens"] if "ens" in sub else pd.Series(0.0, index=d.index)
        dump = sub["dump"] if "dump" in sub else pd.Series(0.0, index=d.index)
        rows.append({"year": y, "zone": z,
                     "unserved_energy_GWh": float(ens.sum()) / 1e3,
                     "unserved_hours": int((ens > 1.0).sum()),
                     "unserved_peak_MW": float(ens.max()),
                     "curtailment_dump_GWh": float(dump.sum()) / 1e3})
    return pd.DataFrame(rows)


_NOTICE = [
    ("PriceModeling — step vi/vii, 20-year projected dispatch", ""),
    ("", ""),
    ("Horizon", "2027-2046, one weather realization (draw 0), 12 modelled bidding zones."),
    ("Scenario", "scenarios.xlsx, 'reference'. Reference year 2019 supplies the window calendar, "
                 "the firm-stack base, the NTC structure and the hydro/must-run curves."),
    ("Weather", "FR demand and RES come from the step-iii/iv engines on a weathergen realization; "
                "neighbours from their reduced-form own-weather response, levelled by TYNDP factors."),
    ("Configuration", "FLEXIBILITY MODULE ON: per-reactor FR nuclear rigidities, and SoC-constrained "
                      "storage (PSP + BESS). DSR ladder 180/250/500 EUR/MWh. Flex block (DSR + H2 "
                      "peakers) bids 180. FR 2050 flexibility = 28 GW: 17.5 BESS + 10.5 DSR/H2, a "
                      "62.5% battery share applied to every zone."),
    ("MODERATED DEMAND", "French demand follows the reduced trajectory: 441.5 (2025) / 474.9 (2030) / "
                         "514.0 (2035) / 538.5 (2040) / 590.1 (2050) TWh, against 460/540/616/680/760 "
                         "before. The model tracks it closely — 473.4 realised in 2030, 535.2 in 2040, "
                         "557.0 in 2046. Neighbour demand paths are UNCHANGED."),
    ("", ""),
    ("TWO PRICE SERIES", ""),
    ("SMC", "The LP balance dual. Internally consistent; congestion rent on it is non-negative "
            "by construction."),
    ("spot", "SMC plus the fitted step-vii markup, clipped to [-500, 4000] EUR/MWh. What a generator "
             "is paid, so revenues lead on it. Where SMC exceeds 4000 the clip bites and spot < SMC — "
             "that happens in 2043 and 2045 only (see below)."),
    ("Congestion rent", "sum_h flow_h x (price_importing - price_exporting), the scarcity rent on the "
                        "NTC constraint. This is the TOTAL rent on the border; in practice it is shared "
                        "between the two TSOs (conventionally 50/50), so RTE's share is about half the "
                        "figure shown. Reported on SMC, where it cannot be negative."),
    ("Revenue", "sum_h generation_h x price_h, from the LP primal — not annual volume x annual mean "
                "price. The difference is the capture-rate effect, which is the whole economic question "
                "for wind and solar. Storage appears as storage_discharge (a sale) and storage_charge "
                "(a purchase, shown negative)."),
    ("", ""),
    ("KNOWN LIMITS — read before using these numbers", ""),
    ("Accuracy", "Backtested on France 2024-25 the model runs MAE 17.6 EUR/MWh, bias +5.1, "
                 "RMSE 23.5 — about 30% of a 59 EUR/MWh level, correlation 0.85-0.87. "
                 "2027-46 error will be larger: 2019-2025 are in-sample for the dispatch too."),
    ("MARKUP / DISPATCH MISMATCH", "reports/markup_model.json was fitted on FLAT-LP backtest SMC. "
                                   "rolling/backtest.py reads the same flexibility flag that is now ON, "
                                   "so the calibration and the dispatch are no longer the same model. "
                                   "The wedge is small (shrink 0.25, mean 1-5 EUR/MWh in most zones) "
                                   "next to what the module moves, but re-deriving the panel under flex "
                                   "is outstanding."),
    ("Storage behaves physically", "FR storage discharges 11.0 -> 27.4 TWh across the horizon from "
                                   "6.0 -> 18.3 GW, i.e. 1375-1811 full-load hours (~1 cycle/day for a "
                                   "4 h fleet) at an 0.82-0.86 round-trip. The flat block it replaced ran "
                                   "2866 FLH on free, unlimited energy. _BESS_PROJ_DURATION_H = 4.0 h is "
                                   "the assumption that makes a stated GW figure mean firm capacity."),
    ("ADEQUACY IS CLOSED IN FRANCE", "Zero unserved energy in every French year, and no hour at the "
                                     "15000 VOLL. On the previous 760 TWh demand path the same "
                                     "configuration shed 68.5 GWh in 2043 and 30.1 in 2045; moderating "
                                     "demand removes both. Across the whole system only DE_LU still sheds "
                                     "load — 612 GWh over 20 YEARS — and it is the one zone whose demand "
                                     "path was not moderated."),
    ("The price ceiling barely binds", "FR max SMC is 180-264 EUR/MWh in 19 of 20 years: the system "
                                       "rarely reaches even the FIRST DSR rung (180). Only 2043 touches "
                                       "500. The scarcity ladder is therefore no longer setting the "
                                       "French price level, which it was on the previous demand path."),
    ("Two scarcity parameters share one value", "The DSR ladder's first rung (180) equals VOM['flex'] "
                                                "(180) exactly, as 300 did before. A sensitivity on "
                                                "either alone leaves ~40% of the affected hours pinned "
                                                "by the other — measured on the old 300 pair. Move both."),
    ("...and two zones are a surplus problem", "AT_SI prices at or below zero in ~90% of hours for a "
                                               "single-digit annual mean; PL_CZ sits well below its "
                                               "neighbours. Both received RES capacity anchors without a "
                                               "matching demand path. Flexibility does not touch this."),
    ("Great Britain markup", "GB carries NO markup: its backtested SMC reached the 15000 VOLL in 89-155 "
                             "hours of every fitted year, so it fails the panel's dispersion gate and "
                             "projects on clipped SMC."),
    ("Missing zones", "IT_SOUTH and the other Italian zones are not modelled; IT_NORTH stands for Italy. "
                      "DK, PL_CZ and AT_SI carry no fitted markup (spot = SMC)."),
    ("Zones, not countries", "DE_LU, PL_CZ and AT_SI are multi-country bidding zones, not states."),
    ("Monte Carlo", "This is ONE draw, not a distribution. Config declares n_draws = 50, and the flex "
                    "capacities were sized from the worst hour of this single draw."),
]


def _style(path: Path):
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path)
    head_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1F3864")
    body_font = Font(name="Arial", size=10)
    for ws in wb.worksheets:
        big = ws.max_row > 5000
        for c in ws[1]:
            c.font = head_font; c.fill = head_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"
        for i, col in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
            hdr = str(col[0].value or "")
            ws.column_dimensions[get_column_letter(i)].width = min(max(11, len(hdr) + 2), 30)
        if big:                                    # 175k-row sheets: skip per-cell styling, it is O(cells)
            continue
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.font = body_font
                if isinstance(c.value, float):
                    c.number_format = "#,##0.00" if abs(c.value) < 1000 else "#,##0"
    wb.save(path)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", default="scratchpad/proj20y")
    ap.add_argument("--out", default="reports/projection_20y")
    ap.add_argument("--hub", default="FR")
    args = ap.parse_args()

    src, out = Path(args.src), Path(args.out)
    out.mkdir(parents=True, exist_ok=True)
    years = sorted(int(p.stem.split("_")[1]) for p in src.glob("prices_*.parquet"))
    if not years:
        print(f"no prices_*.parquet in {src}", file=sys.stderr)
        return 1
    print(f"[deliverables] {len(years)} years: {years[0]}-{years[-1]}", flush=True)

    spot_all, rev, cong, neg, psum, sysr = [], [], [], [], [], []
    for y in years:
        smc, spot, d, fl = _load_year(src, y)
        s = spot.copy()
        s.insert(0, "FR_SMC", smc["FR"])
        spot_all.append(s)
        psum.append(price_summary(y, smc, spot))
        neg.append(negative_hours(y, smc, spot))
        if d is not None:
            rev.append(revenues(y, smc, spot, d))
            sysr.append(system_table(y, d))
        if fl is not None:
            cong.append(congestion(y, smc, spot, fl, hub=args.hub))
        print(f"  {y} ok", flush=True)

    prices = pd.concat(spot_all).sort_index()
    prices.index.name = "timestamp_utc"
    cols = [args.hub] + [c for c in prices.columns if c not in (args.hub, "FR_SMC")] + ["FR_SMC"]
    prices = prices[cols].round(2)
    # Excel has no tz-aware datetime. The index is UTC and the column name says so; dropping the offset
    # here rather than converting keeps every hour on the same clock across the horizon's DST boundaries.
    prices.index = prices.index.tz_localize(None)

    tables = {
        "Revenue by tech": pd.concat(rev, ignore_index=True) if rev else pd.DataFrame(),
        "Congestion rent": pd.concat(cong, ignore_index=True) if cong else pd.DataFrame(),
        "Negative & zero hours": pd.concat(neg, ignore_index=True),
        "Price summary": pd.concat(psum, ignore_index=True),
        "System adequacy": pd.concat(sysr, ignore_index=True) if sysr else pd.DataFrame(),
    }

    for name, df in tables.items():
        df.to_csv(out / (name.lower().replace(" & ", "_").replace(" ", "_") + ".csv"), index=False)
    prices.to_csv(out / "hourly_spot_prices.csv")

    notice = pd.DataFrame(_NOTICE, columns=["Field", "Value"])

    # year x category views of the same numbers — the long tables are the record, these are readable
    pivots = {}
    rv = tables["Revenue by tech"]
    if not rv.empty:
        pivots[f"{args.hub} revenue MEUR"] = rv[rv["zone"] == args.hub].pivot_table(
            index="year", columns="tech", values="revenue_MEUR_spot", aggfunc="sum").round(1)
        pivots[f"{args.hub} capture EURperMWh"] = rv[rv["zone"] == args.hub].pivot_table(
            index="year", columns="tech", values="capture_price_spot", aggfunc="sum").round(2)
        pivots["Revenue by zone MEUR"] = rv.pivot_table(
            index="year", columns="zone", values="revenue_MEUR_spot", aggfunc="sum").round(1)
    if not tables["Congestion rent"].empty:
        pivots["Congestion rent MEUR"] = tables["Congestion rent"].pivot_table(
            index="year", columns="border", values="congestion_rent_MEUR_smc", aggfunc="sum").round(1)
    pivots["Hours <= 0"] = tables["Negative & zero hours"].pivot_table(
        index="year", columns="zone", values="hours_le_zero_spot", aggfunc="sum")
    pivots["Mean spot"] = tables["Price summary"].pivot_table(
        index="year", columns="zone", values="mean_spot", aggfunc="sum").round(2)
    # a named columns axis makes pandas write an extra header row, which breaks the frozen pane
    pivots = {k: v.rename_axis(columns=None) for k, v in pivots.items()}
    for name, df in pivots.items():
        df.to_csv(out / ("pivot_" + name.lower().replace(" ", "_").replace("<=", "le") + ".csv"))

    f2 = out / "projection_20y_analysis.xlsx"
    with pd.ExcelWriter(f2, engine="openpyxl") as xl:
        notice.to_excel(xl, sheet_name="Notice", index=False)
        for name, df in pivots.items():
            df.to_excel(xl, sheet_name=name[:31])
        for name, df in tables.items():
            df.round(3).to_excel(xl, sheet_name=name[:31], index=False)
    _style(f2)
    print(f"[deliverables] wrote {f2}", flush=True)

    f1 = out / "projection_20y_spot_prices.xlsx"
    with pd.ExcelWriter(f1, engine="openpyxl") as xl:
        notice.to_excel(xl, sheet_name="Notice", index=False)
        tables["Price summary"].round(2).to_excel(xl, sheet_name="Yearly summary", index=False)
        prices.to_excel(xl, sheet_name="Hourly spot")
    _style(f1)
    print(f"[deliverables] wrote {f1}  ({len(prices):,} hours x {prices.shape[1]} columns)", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
