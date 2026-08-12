"""Fill PT's rows in `dispatch_tyndp`. Portugal is a SCORED zone that had no trajectory at all.

PT is scored by both harnesses (`scripts/gate_multiyear.py`, `scripts/projection_backcast.py`) yet carried
no TYNDP row, so every projected year ran it on the generic CAGR: RES +4.5 %/yr and demand +0.8 %/yr. That
compounds Portuguese solar to 3.2 GW by 2046 against a national target of 20.8 GW by 2030 — off by a factor
of six, in a zone whose price the model reports and grades.

SOURCE: revised **PNEC 2030** (Plano Nacional Energia e Clima), the draft submitted for public consultation
on 22 July 2024 — the same national-plan basis `8be1a62` used for FR/DE/ES/IT/BE/NL/CH. Figures quoted from
the Pérez-Llorca legal briefing of 23 July 2024 summarising that draft:

    "the new target for the installed capacity of solar technology has been increased from 20.4 GW
     to 20.8 GW"
    "the targets for installed capacity of both onshore and offshore wind have been maintained —
     10.4 GW and 2 GW, respectively"
    "it is expected that Portugal will increase its figures of 50 TWh in 2024 to 90 TWh in 2030"

<https://www.perezllorca.com/wp-content/uploads/2024/07/Legal-Briefing-Revised-Portuguese-National-Energy-and-Climate-Plan-2021-2030-esg.pdf>

THE 2019 SOLAR ANCHOR IS NOT ENTSO-E'S, AND THAT MATTERS. `entsoe_installed_capacity` declares 324 MW of
Portuguese solar for 2019, but the metered fleet PEAKED at 517 MW that year — it out-produced its own
declared nameplate by 60 %, which is impossible — and generated 1.062 TWh, implying a capacity factor of
0.37 against a declaration of 324 MW. IRENA/Ember put the end-2019 fleet at **970 MW**, giving a plausible
CF of ~0.14. The declaration is a ~3x under-report, the same defect `TYNDP_SOURCES.md` records for Swiss
solar, and it is load-bearing here: the `res` factor scales OBSERVED 2019 generation, which was produced by
the true fleet, so anchoring on 324 MW would inflate the growth ratio threefold (x64 instead of x21).

PT WIND'S 2019 ANCHOR IS ENTSO-E'S, cross-checked and sound: 5127 MW declared against a metered peak of
4632 MW (90 % of nameplate), which is what a healthy declaration looks like.

DEMAND ANCHORS IN-HOUSE, and the definitions were verified to match rather than assumed. The lake's PT load
totals 50.35 TWh in 2019 and 51.41 TWh in 2024, against the PNEC's "50 TWh in 2024" — so the plan's
consumption figure and the model's load series measure the same quantity, and the 2019 anchor can come from
the lake while the 2030 target comes from the plan.

ANCHORS STOP AT 2030, DELIBERATELY. RNC2050 (Roteiro para a Neutralidade Carbónica 2050, 2019) is the
obvious source for 2040/2050 and is SUPERSEDED: it envisages ~13 GW of solar by 2050 where PNEC 2030 already
targets 20.8 GW by 2030, so splicing it in would make Portuguese capacity DECLINE after 2030. Clamping flat
at the 2030 level is wrong too, but it is wrong by far less than either the CAGR (3.2 GW by 2046) or a
decline. Replacing this with a post-PNEC long-term source is the follow-up.

Run from dispatch_model/:  python -X utf8 scripts/gen_tyndp_pt.py [--write]
"""
from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dispatch_model.config import load_config                              # noqa: E402

#: INTERMEDIATE ANCHORS ARE NOT OPTIONAL HERE, and the first version of this script omitted them. `_interp`
#: is LINEAR, so a 2019 -> 2030 pair spreads PNEC's growth uniformly — but the plan is back-loaded, with most
#: of the 20.8 GW arriving after 2025. Two endpoints put Portuguese solar at 9.98 GW in 2024 against an
#: actual ~5.4 GW, and the backcast measured the cost: PT projected 39.1 EUR/MWh against 63.5 observed, its
#: layer going -4.3 -> -29.9 and dragging ES from -4.0 to -16.5 through the Iberian border. The endpoints
#: were right; the SHAPE between them was invented.
#:
#: The intermediate solar anchors come from the lake's own series RESCALED by the measured under-report
#: factor. ENTSO-E under-declares Portuguese solar in every year (metered peak exceeds declared nameplate
#: 1.17-1.60x), but it under-declares it CONSISTENTLY: lake 0.324 vs IRENA 0.970 in 2019 is x2.99, and lake
#: 1.332 vs IRENA 4.040 in 2023 is x3.03. A stable factor means the level is wrong while the trajectory is
#: sound, so lake x3.0 is a defensible true-level series — calibrated on two independent IRENA points, not
#: assumed.
ROWS = [
    # (variable, year, value, provenance)
    ("cap_solar_gw", 2019, 0.970, "IRENA/Ember end-2019; ENTSO-E's 0.324 rejected (see docstring)"),
    ("cap_solar_gw", 2023, 4.040, "IRENA/Ember end-2023; lake 1.332 x 3.03 agrees"),
    ("cap_solar_gw", 2024, 5.430, "lake 1.811 x 2.998, the factor calibrated on 2019 and 2023"),
    ("cap_solar_gw", 2030, 20.800, "PNEC 2030 revised draft, solar installed capacity target"),
    ("cap_wind_gw", 2019, 5.127, "entsoe_installed_capacity 2019 (peak 4.632 GW = 90 % of nameplate)"),
    ("cap_wind_gw", 2024, 5.333, "entsoe_installed_capacity 2024 (peak/declared 0.94 — declaration sound)"),
    ("cap_wind_gw", 2030, 12.400, "PNEC 2030 revised: 10.4 GW onshore + 2.0 GW offshore"),
    ("demand_twh", 2019, 50.350, "entsoe_load 2019 total (lake)"),
    ("demand_twh", 2024, 50.000, "PNEC 2030 revised, stated 2024 level; lake reads 51.41 -> definitions agree"),
    ("demand_twh", 2030, 90.000, "PNEC 2030 revised, 2030 consumption"),
]


def build_rows() -> list[dict]:
    return [{"zone": "PT", "variable": v, "year": y, "value": x} for v, y, x, _ in ROWS]


if __name__ == "__main__":
    cfg = load_config("config.yaml")
    wb_path = cfg.resolve(cfg.section("assumptions")["workbook"])
    rows = build_rows()
    print(f"{len(rows)} PT rows for `dispatch_tyndp`\n")
    for v, y, x, src in ROWS:
        print(f"  {v:<14} {y}  {x:>8.3f}   {src}")

    df = pd.DataFrame(rows)
    print("\n  implied factors vs the 2019 reference year:")
    for var in ("cap_solar_gw", "cap_wind_gw", "demand_twh"):
        g = df[df.variable == var].sort_values("year")
        base = float(g[g.year == 2019].value.iloc[0])
        print(f"    {var:<14} " + "  ".join(
            f"{int(r.year)}=x{r.value / base:.2f}" for r in g.itertuples() if r.year != 2019))
    res19 = 0.970 + 5.127
    res30 = 20.800 + 12.400
    print(f"\n    RES total (what `_RES_VARS` sums): {res19:.3f} -> {res30:.3f} GW = x{res30 / res19:.2f}")
    print(f"    demand:                            50.35 -> 90.00 TWh = x{90.0 / 50.35:.2f}")
    print("    RES outgrows demand ~3x, which is what a 93 %-renewable-electricity target means.")

    if "--write" not in sys.argv:
        print("\n(dry run — pass --write to update the workbook)")
        raise SystemExit

    from openpyxl import load_workbook                                     # noqa: E402
    wb = load_workbook(wb_path)
    ws = wb["dispatch_tyndp"]
    hdr = {c.value: i + 1 for i, c in enumerate(ws[1])}
    zc, vc, yc, valc = hdr["zone"], hdr["variable"], hdr["year"], hdr["value"]
    index = {(ws.cell(r, zc).value, ws.cell(r, vc).value, int(ws.cell(r, yc).value)): r
             for r in range(2, ws.max_row + 1) if ws.cell(r, yc).value is not None}
    n_e = n_a = 0
    for r_ in rows:
        key = (r_["zone"], r_["variable"], r_["year"])
        if key in index:
            ws.cell(index[key], valc).value = r_["value"]; n_e += 1
        else:
            rr = ws.max_row + 1
            ws.cell(rr, zc).value = r_["zone"]; ws.cell(rr, vc).value = r_["variable"]
            ws.cell(rr, yc).value = r_["year"]; ws.cell(rr, valc).value = r_["value"]
            index[key] = rr; n_a += 1
    wb.save(wb_path)
    print(f"\nworkbook updated: {n_e} replaced, {n_a} added")
