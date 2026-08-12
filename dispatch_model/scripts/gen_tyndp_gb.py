"""Fill GB's rows in `dispatch_tyndp` — the gap that costs the backcast a -10.7 EUR/MWh layer.

GB was promoted to a modelled zone without capacity trajectories, so `tyndp.report_coverage` has printed
`GB: cap_solar_gw=missing, cap_wind_gw=missing, demand_twh=missing` on every projection run since, and the
projection fell back to the generic CAGR: RES +4.5 %/yr against demand +0.8 %/yr, i.e. RES x1.25 vs demand
x1.04 over 2019-2024. That over-grows surplus exactly in the direction measured — GB's projection arm
prints 60.2 EUR/MWh against 84.5 observed, the worst zone in the harness, of which -10.7 is layer.

WHY NOT TYNDP. Every other zone's rows come from TYNDP, and GB cannot: it left ENTSO-E after Brexit. TYNDP
2024 still models `UK00` in its reference GRID (which is where this repo's GB border capacities come from),
but its published scenario figures are EU27 aggregates plus low/high ranges at 2040/2050 — there is no
per-country UK capacity trajectory with a usable early anchor. So GB uses its own system operator's
scenarios, exactly as CH's solar came from BFE/Swissolar rather than from ENTSO-E.

SOURCE: NESO *Future Energy Scenarios 2024*, Data Workbook, sheets **ES.14** (solar capacity), **ES.13**
(onshore wind), **ES.12** (offshore wind), **DB.ED1** (electricity demand), scenario **Holistic
Transition** — NESO's central pathway, the closest analogue to the TYNDP National Trends+ basis the other
zones use. The Counterfactual (no further policy) and Electric Engagement (high electrification) rows
bracket it and are recorded below for anyone wanting a sensitivity.
<https://www.neso.energy/publications/future-energy-scenarios-fes/fes-documents> (Data Workbook 2024)

THE 2019 ANCHOR IS MEASURED IN-HOUSE, NOT TAKEN FROM FES. FES 2024's capacity series begin at 2023, and
`tyndp_factors` computes target/reference against the projection's 2019 reference year — an anchor that
starts after it clamps, which `tyndp.coverage` correctly calls the dangerous class. But GB's 2019 installed
capacity is already in the lake: ENTSO-E published it before Brexit, in the same
`entsoe_installed_capacity` table, on the same nameplate definition every other zone's anchor uses. So the
ratio is anchored on the same measurement basis as the rest of the workbook rather than on a foreign one.

    entsoe_installed_capacity, series_key GB, 2019
        Solar          13 346 MW  -> cap_solar_gw 13.346
        Wind Onshore   12 638 MW  \\
        Wind Offshore   9 379 MW  /  -> cap_wind_gw 22.017   (onshore + offshore, per tyndp._RES_VARS)

    continuity check against FES 2023: solar 13.35 -> 15.14, wind 22.02 -> 28.41. Both consistent with
    four years of build-out, so the two sources join without a step.

DEMAND IS DELIBERATELY LEFT CLAMPED, and this is the one compromise here. FES's demand series starts at
2022 and is *national* demand (343 TWh in 2022); the lake's GB load is Elexon ITSDO, demand at the
TRANSMISSION boundary (~249 TWh), which is a different quantity by ~95 TWh of embedded generation and
losses. Splicing them would put a 38 % step into a ratio that is supposed to measure structural growth.
FES 2022 is therefore the earliest demand anchor and `_interp(2019)` clamps to it. The cost of that clamp
is small where it matters: GB demand moved only a few per cent between 2019 and 2022, against RES which
roughly doubled — so the clamp misprices the smaller term while the larger one is properly anchored.

Run from dispatch_model/:  python -X utf8 scripts/gen_tyndp_gb.py [--write]
"""
from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dispatch_model.config import load_config                              # noqa: E402

#: NESO FES 2024 Data Workbook, Holistic Transition. GW.
FES_SOLAR = {2023: 15.136, 2025: 21.793, 2030: 40.471, 2035: 69.189,
             2040: 85.213, 2045: 100.922, 2050: 107.886}
FES_WIND_ON = {2023: 13.690, 2025: 16.799, 2030: 27.369, 2035: 31.246,
               2040: 34.697, 2045: 38.888, 2050: 40.213}
FES_WIND_OFF = {2023: 14.723, 2025: 23.204, 2030: 53.556, 2035: 88.570,
                2040: 98.474, 2045: 98.474, 2050: 100.474}
#: DB.ED1, Holistic Transition, TWh. Starts 2022 — see the docstring on the clamp.
FES_DEMAND = {2022: 343.4, 2025: 336.3, 2030: 376.2, 2035: 455.9,
              2040: 544.4, 2045: 629.3, 2050: 662.9}

#: measured, `entsoe_installed_capacity` series_key GB year 2019 (pre-Brexit ENTSO-E submission)
LAKE_2019 = {"cap_solar_gw": 13.346, "cap_wind_gw": 22.017}


def build_rows() -> list[dict]:
    rows = [{"zone": "GB", "variable": v, "year": 2019, "value": x} for v, x in LAKE_2019.items()]
    for y, v in FES_SOLAR.items():
        rows.append({"zone": "GB", "variable": "cap_solar_gw", "year": y, "value": round(v, 3)})
    for y in FES_WIND_ON:
        rows.append({"zone": "GB", "variable": "cap_wind_gw", "year": y,
                     "value": round(FES_WIND_ON[y] + FES_WIND_OFF[y], 3)})
    for y, v in FES_DEMAND.items():
        rows.append({"zone": "GB", "variable": "demand_twh", "year": y, "value": round(v, 1)})
    return rows


if __name__ == "__main__":
    cfg = load_config("config.yaml")
    wb_path = cfg.resolve(cfg.section("assumptions")["workbook"])
    rows = build_rows()
    df = pd.DataFrame(rows)
    print(f"{len(df)} GB rows for `dispatch_tyndp`\n")
    for var, g in df.groupby("variable"):
        g = g.sort_values("year")
        print(f"  {var:<15} " + "  ".join(f"{int(r.year)}={r.value:g}" for r in g.itertuples()))
    print("\n  implied ratios vs the 2019 reference year (what `tyndp_factors` will apply):")
    for var in ("cap_solar_gw", "cap_wind_gw"):
        g = df[df.variable == var].sort_values("year")
        base = float(g[g.year == 2019].value.iloc[0])
        print(f"    {var:<15} " + "  ".join(
            f"{int(r.year)}=x{r.value / base:.2f}" for r in g.itertuples() if r.year != 2019))
    g = df[df.variable == "demand_twh"].sort_values("year")
    base = float(g.value.iloc[0])
    print(f"    {'demand_twh':<15} " + "  ".join(
        f"{int(r.year)}=x{r.value / base:.2f}" for r in g.itertuples()) + "   (2019 clamps to 2022)")

    if "--write" not in sys.argv:
        print("\n(dry run — pass --write to update the workbook)")
        raise SystemExit

    from openpyxl import load_workbook                                     # noqa: E402
    wb = load_workbook(wb_path)
    ws = wb["dispatch_tyndp"]
    hdr = {c.value: i + 1 for i, c in enumerate(ws[1])}
    zc, vc, yc, valc = hdr["zone"], hdr["variable"], hdr["year"], hdr["value"]
    index = {(ws.cell(r, zc).value, ws.cell(r, vc).value, int(ws.cell(r, yc).value)): r
             for r in range(2, ws.max_row + 1) if ws.cell(r, yc).value is not None}
    n_e = n_a = 0
    for r_ in rows:
        key = (r_["zone"], r_["variable"], r_["year"])
        if key in index:
            ws.cell(index[key], valc).value = r_["value"]; n_e += 1
        else:
            rr = ws.max_row + 1
            ws.cell(rr, zc).value = r_["zone"]; ws.cell(rr, vc).value = r_["variable"]
            ws.cell(rr, yc).value = r_["year"]; ws.cell(rr, valc).value = r_["value"]
            index[key] = rr; n_a += 1
    wb.save(wb_path)
    print(f"\nworkbook updated: {n_e} replaced, {n_a} added")
