"""Pre-fill `dispatch_tyndp`'s `cap_nuclear_gw` rows from the policy rule (stacks.nuclear_fleet).

Phase-out countries (DE, ES) retire on their committed schedules; the others run to a 60-year lifetime.
France is derived reactor-by-reactor from `avail_fleet_registry` commissioning years joined to the real
capacities in the FR stack (the registry's own `capacity_mw` column is not capacity — Belleville 1 shows
6998 for a 1310 MW reactor — so only the years are taken from it).

Values are written as a DEFAULT: existing rows are replaced, and the user can edit any cell afterwards.
Run from dispatch_model/:  python -X utf8 scripts/gen_nuclear_trajectory.py [--write]
"""
from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dispatch_model.config import load_config                    # noqa: E402
from dispatch_model.stacks import nuclear_fleet as nf            # noqa: E402
from dispatch_model.stacks.fr_stack import build_fr_stack        # noqa: E402

# 2019 is the projection's REFERENCE year: `tyndp_factors` computes target/reference, so without a
# reference-year row the ratio is undefined and the tech escapes scaling entirely.
YEARS = (2019, 2025, 2030, 2040, 2050)
ZONES = ("FR", "BE", "CH", "ES", "DE_LU", "NL")

cfg = load_config("config.yaml")
wb_path = cfg.resolve(cfg.section("assumptions")["workbook"])

# FR reactor-by-reactor: real capacity from the stack, commissioning year from the fleet registry
stack = build_fr_stack(cfg)
stack = stack[stack["tech"] == "nuclear"][["name", "capacity_mw"]]
reg = pd.read_excel(wb_path, sheet_name="avail_fleet_registry")
reg = reg[reg["technology"] == "nuclear"][["name", "commissioning_year"]]
fr = stack.merge(reg, on="name", how="left")
fr_units, unknown = [], []
for r in fr.itertuples():
    com, closed = nf.FR_MISSING.get(str(r.name), (None, None))
    if r.commissioning_year == r.commissioning_year:            # registry value present
        com = int(r.commissioning_year)
    if com is None:
        unknown.append(str(r.name))
        continue
    fr_units.append((float(r.capacity_mw), int(com), closed))
live = [(m, c, x) for m, c, x in fr_units if x is None]
print(f"FR: {len(fr_units)}/{len(stack)} reactors dated "
      f"({sum(m for m, _, _ in fr_units) / 1e3:.1f} GW nameplate, "
      f"{sum(m for m, _, _ in live) / 1e3:.1f} GW still open today)")
if unknown:
    print(f"    WITHOUT a commissioning year (excluded): {unknown}")

# EPR2 new build — a DEDICATED scenario input (`dispatch_nuclear_newbuild`), created and pre-filled
# from EDF's officially retained schedule on first run, then read back so the user's edits win.
try:
    nb = pd.read_excel(wb_path, sheet_name="dispatch_nuclear_newbuild")
    print(f"\nEPR2 sheet read: {len(nb)} units")
except (ValueError, KeyError):
    nb = pd.DataFrame([{"zone": z, "unit": u, "capacity_mw": mw,
                        "commissioning_year": y, "scenario": "reference"}
                       for z, u, mw, y in nf.EPR2])
    print(f"\nEPR2 sheet absent — will create it with EDF's retained dates ({len(nb)} units)")
newbuild = [(str(r.zone), float(r.capacity_mw), int(r.commissioning_year)) for r in nb.itertuples()]

rows = nf.trajectory(ZONES, YEARS, fr_units=fr_units, newbuild=newbuild)
tab = pd.DataFrame(rows).pivot_table(index="zone", columns="year", values="value")
print("\ncap_nuclear_gw (policy default):")
print(tab.to_string())

if "--write" not in sys.argv:
    print("\n(dry run — pass --write to update the workbook)")
    raise SystemExit

from openpyxl import load_workbook                               # noqa: E402

wb = load_workbook(wb_path)
if "dispatch_nuclear_newbuild" not in wb.sheetnames:        # create the dedicated scenario input
    ns = wb.create_sheet("dispatch_nuclear_newbuild")
    ns.append(["zone", "unit", "capacity_mw", "commissioning_year", "scenario"])
    for z, u, mw, y in nf.EPR2:
        ns.append([z, u, mw, y, "reference"])
    print("created sheet dispatch_nuclear_newbuild (EDF retained EPR2 dates, editable)")
ws = wb["dispatch_tyndp"]
hdr = {c.value: i + 1 for i, c in enumerate(ws[1])}
zc, vc, yc, valc = hdr["zone"], hdr["variable"], hdr["year"], hdr["value"]
existing = {}
for r in range(2, ws.max_row + 1):
    if ws.cell(r, vc).value == "cap_nuclear_gw":
        existing[(ws.cell(r, zc).value, int(ws.cell(r, yc).value))] = r
written, added = 0, 0
for row in rows:
    key = (row["zone"], row["year"])
    if key in existing:
        ws.cell(existing[key], valc).value = row["value"]
        written += 1
    else:
        r = ws.max_row + 1
        ws.cell(r, zc).value = row["zone"]
        ws.cell(r, vc).value = "cap_nuclear_gw"
        ws.cell(r, yc).value = row["year"]
        ws.cell(r, valc).value = row["value"]
        added += 1
wb.save(wb_path)
print(f"\nworkbook updated: {written} rows replaced, {added} added")
