"""Add the REFERENCE-YEAR (2019) baseline rows to `dispatch_tyndp`.

`tyndp_factors` computes `interp(target) / interp(reference_year)`. The tab's anchors started at 2025,
so `interp(2019)` extrapolated to the 2025 value and every factor was really *target ÷ 2025* — applied
to a 2019 stack and weather shape. The whole 2019→2025 structural change was silently dropped:
measured, every projected year realised `actual2019 × target/2025` of RES instead of `target`
(FR solar 41 %, ES solar 23 %), and a 2024 target got factors of exactly 1.0 — the projection replayed
2019 unchanged (`scratchpad/projection_backcast_2024.py`).

This writes the missing 2019 anchor from ACTUAL data, so the factor becomes target ÷ actual-2019 applied
to the actual-2019 stack. Rules:
  * only for a variable the zone ALREADY anchors in the future — inventing a lone 2019 row would replace
    the deliberate CAGR fallback with a constant 1.0;
  * `cap_flex_gw` is excluded: `flex_capacity_mw` reads it as an ABSOLUTE level, and its own docstring
    records that there is no reference-year flex fleet to scale from;
  * `cap_nuclear_gw` is left to `gen_nuclear_trajectory.py`, which already emits a policy-derived 2019.
Run from dispatch_model/:  python -X utf8 scripts/gen_tyndp_baseline.py [--write]
"""
from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dispatch_model.config import load_config                              # noqa: E402
from dispatch_model.io.entsoe_hist import load_demand_hist, load_installed_capacity   # noqa: E402
from dispatch_model.neighbours.blocks import constituents                  # noqa: E402

REF_YEAR = 2019
SKIP = {"cap_flex_gw", "cap_nuclear_gw"}

# ---- cap_hydro_gw: the definition, and why one tech set cannot baseline every zone ----------------
# THE DEFINITION. `cap_hydro_gw` means RESERVOIR + RUN-OF-RIVER, EXCLUDING pumped storage. That is what
# `tyndp._CAP_VAR` already encodes (`hydro_psp` maps to its own `cap_psp_gw`), and the dispatch forces
# it: PSP is not in the hydro merit order at all — `flexibility.storage` sizes the storage LP from the
# zone's measured `hydro_psp` stack capacity. A PSP-inclusive factor would scale a stack that does not
# contain the PSP it is counting.
#
# THE SHEET DOES NOT FOLLOW IT, and not uniformly. Measured against 2019 installed capacity:
#
#     zone      res+ror   +psp    2025 anchor   entered on
#     FR          19.23   24.26          26.0   PSP-INCLUSIVE  (deviates)
#     CH           6.05   12.69          15.0   PSP-INCLUSIVE  (deviates)
#     DE_LU        5.32   14.74           5.0   res+ror        (correct)
#     ES          20.30   25.95          20.0   res+ror        (correct)
#     IT_NORTH     0.00    0.00          13.0   undeterminable (no capacity rows at all)
#
# This is why the previous attempt was abandoned: it applied ONE tech set to every zone, so it was right
# for half of them and wrong for the other half — PSP-inclusive gave DE 0.34 (deleting two-thirds of
# German hydro), PSP-exclusive gave FR 1.41 (inventing 41 % growth).
#
# THE FIX. `tyndp_factors` consumes a RATIO within a zone, so a constant definitional offset CANCELS —
# provided the reference-year baseline is measured on the same basis as that zone's own future anchors.
# So detect the basis per zone (which of the two candidate sums the first future anchor is closer to)
# and baseline on it. Nothing is invented and no scenario value is edited.
#
# BE HONEST ABOUT THE SIZE OF THIS. The anchors are near-flat — FR ×1.038 and CH ×1.067 over 2025-2050,
# DE_LU / ES / IT_NORTH exactly ×1.000 — so the clamped 1.0 was already almost right. Fixing it is worth
# at most ~4 % (FR) and ~7 % (CH) at 2050 and near zero elsewhere. It is done because a permanently
# "clamped" variable masks the real gaps in the coverage report, not because it moves prices much.
#
# The DEFINITIVE fix is still to normalise the sheet: restate FR and CH on the res+ror basis and put
# their PSP in `cap_psp_gw`. That edits scenario values, and CH needs a call on how PSP evolves (Swiss
# PSP is NOT static — Nant de Drance added ~0.9 GW in 2022), so it is left to the workbook owner.
HYDRO_BASES = {
    "res+ror": ("hydro_reservoir", "hydro_ror"),
    "res+ror+psp": ("hydro_reservoir", "hydro_ror", "hydro_psp"),
}
#: reject a basis match worse than this (relative) — the anchor is then not explained by either
#: definition and a baseline would be a guess. Catches IT_NORTH, which has no capacity rows at all.
HYDRO_BASIS_TOL = 0.35
#: TYNDP variable → the installed-capacity techs that make it up
VAR2TECH = {
    "cap_solar_gw": ("solar",), "cap_wind_gw": ("wind_onshore", "wind_offshore"),
    "cap_gas_gw": ("gas",), "cap_coal_gw": ("coal",), "cap_lignite_gw": ("lignite",),
    "cap_oil_gw": ("oil",), "cap_biomass_gw": ("biomass",),
    # `cap_hydro_gw` counts PSP: no zone defines `cap_psp_gw`, and the anchors are plainly TOTAL hydro
    # (FR 26 GW ≈ France's 25.7 GW of reservoir + ROR + pumped storage). Baselining on reservoir+ROR
    # alone would read 19.2 GW for FR and manufacture 41 % hydro growth by 2040 out of a definitional
    # mismatch. The baseline must be measured the way the anchor is defined.
    "cap_hydro_gw": ("hydro_reservoir", "hydro_ror", "hydro_psp"), "cap_psp_gw": ("hydro_psp",),
}

cfg = load_config("config.yaml")
wb_path = cfg.resolve(cfg.section("assumptions")["workbook"])
sheet = pd.read_excel(wb_path, sheet_name="dispatch_tyndp")


def hydro_basis(zone: str, first_anchor: float) -> tuple[str, float] | None:
    """→ (basis name, GW) for the definition this zone's OWN anchors were entered on, or None.

    Both candidate sums are measured and the one closest to the zone's first future anchor wins. Closest
    in RELATIVE terms: an absolute comparison would favour the larger sum for big-hydro zones regardless
    of definition. Returns None when neither is within `HYDRO_BASIS_TOL`, i.e. the anchor is explained by
    neither definition and any baseline would be a guess.
    """
    best = None
    for name, techs in HYDRO_BASES.items():
        gw = measured_gw(zone, techs)
        if gw is None or gw <= 0:
            continue
        err = abs(first_anchor - gw) / first_anchor
        if best is None or err < best[2]:
            best = (name, gw, err)
    if best is None or best[2] > HYDRO_BASIS_TOL:
        return None
    return best[0], best[1]


def measured_gw(zone: str, techs: tuple[str, ...]) -> float | None:
    tot = 0.0
    for z in constituents(zone):
        try:
            cap = load_installed_capacity(cfg, z, REF_YEAR) or {}
        except Exception:                                                  # noqa: BLE001
            continue
        tot += sum(float(cap.get(t, 0.0)) for t in techs)
    return round(tot / 1e3, 3) if tot > 0 else None


def actual(zone: str, variable: str) -> float | None:
    zs = constituents(zone)
    if variable == "demand_twh":
        try:
            d = load_demand_hist(cfg, REF_YEAR, zones=zs)
        except Exception:                                                  # noqa: BLE001
            return None
        return round(float(d["load_mw"].sum()) / 1e6, 1) if not d.empty else None
    techs = VAR2TECH.get(variable)
    if not techs:
        return None
    tot = 0.0
    for z in zs:
        try:
            cap = load_installed_capacity(cfg, z, REF_YEAR) or {}
        except Exception:                                                  # noqa: BLE001
            continue
        tot += sum(float(cap.get(t, 0.0)) for t in techs)
    return round(tot / 1e3, 3) if tot > 0 else None


rows, skipped, bases = [], [], []
for (zone, variable), g in sheet.groupby(["zone", "variable"]):
    if variable in SKIP:
        continue
    if not (g["year"] > REF_YEAR).any():                                   # no future anchor → leave alone
        continue
    future = g[g["year"] > REF_YEAR].sort_values("year")
    if variable == "cap_hydro_gw":              # per-zone definition — see the HYDRO_BASES block above
        det = hydro_basis(str(zone), float(future.iloc[0].value))
        if det is None:
            skipped.append(f"{zone}/{variable} (anchor {future.iloc[0].value:g} GW matches neither "
                           f"res+ror nor res+ror+psp within {HYDRO_BASIS_TOL:.0%})")
            continue
        basis, v = det
        a0, a1 = float(future.iloc[0].value), float(future.iloc[-1].value)
        span, gap = a1 / a0 - 1.0, (a0 - v) / v
        # LEVEL GAP — REPORTED, NOT GATED. The anchors are near-flat while ENTSO-E's measured stock and
        # the scenario's own first anchor differ by more than the whole trajectory, so the baseline
        # injects that difference as apparent growth in the first projected years (FR +8.9 %, CH +37 %
        # once restated onto res+ror, against 25-year trajectories of +4.8 % and +12.0 %). This was put
        # to the workbook owner, who accepted the error source explicitly — so the baseline is written
        # and the gap is printed on every run rather than silently swallowed. The remaining fix is on
        # the ANCHOR side: reconcile the scenario levels with the measured stock (ENTSO-E reports 19.2 GW
        # of FR reservoir+ROR and only 6.1 GW for CH, which under-counts Swiss small hydro badly).
        bases.append(f"{zone}: {basis} = {v:g} GW | anchors {a0:g}→{a1:g} span {span:+.1%} | "
                     f"level gap {gap:+.1%}"
                     + ("  ⚠ gap exceeds the trajectory" if abs(gap) > abs(span) else ""))
    else:
        v = actual(str(zone), str(variable))
        if v is None:
            skipped.append(f"{zone}/{variable}")
            continue
    rows.append({"zone": zone, "variable": variable, "year": REF_YEAR, "value": v,
                 "first_anchor": f"{int(future.iloc[0].year)}={future.iloc[0].value:g}"})

df = pd.DataFrame(rows)
print(f"{REF_YEAR} baselines derived from actuals ({len(df)} rows):")
for z, g in df.groupby("zone"):
    bits = [f"{r.variable.replace('cap_', '').replace('_gw', ''):<8}{r.value:>8.1f}  (vs {r.first_anchor})"
            for r in g.itertuples()]
    print(f"  {z}:")
    for b in bits:
        print(f"      {b}")
if bases:
    print("\ncap_hydro_gw — definition detected per zone from that zone's own anchors")
    print("(the model's definition is res+ror EXCLUDING psp; a zone marked res+ror+psp deviates from it,")
    print(" but the RATIO still cancels the offset once the baseline uses the same basis):")
    for b in bases:
        print(f"      {b}")
if skipped:
    print(f"\nno actual available (left to the existing behaviour): {skipped}")

if "--write" not in sys.argv:
    print("\n(dry run — pass --write to update the workbook)")
    raise SystemExit

from openpyxl import load_workbook                                         # noqa: E402

wb = load_workbook(wb_path)
ws = wb["dispatch_tyndp"]
hdr = {c.value: i + 1 for i, c in enumerate(ws[1])}
zc, vc, yc, valc = hdr["zone"], hdr["variable"], hdr["year"], hdr["value"]
existing = {(ws.cell(r, zc).value, ws.cell(r, vc).value, int(ws.cell(r, yc).value)): r
            for r in range(2, ws.max_row + 1) if ws.cell(r, yc).value is not None}
repl = add = 0
for r_ in rows:
    key = (r_["zone"], r_["variable"], REF_YEAR)
    if key in existing:
        ws.cell(existing[key], valc).value = r_["value"]
        repl += 1
    else:
        rr = ws.max_row + 1
        ws.cell(rr, zc).value = r_["zone"]
        ws.cell(rr, vc).value = r_["variable"]
        ws.cell(rr, yc).value = REF_YEAR
        ws.cell(rr, valc).value = r_["value"]
        add += 1
wb.save(wb_path)
print(f"\nworkbook updated: {repl} replaced, {add} added")
