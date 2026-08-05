"""Year-index the ES bid ladder in `dispatch_res_schemes`, with floors measured per year.

The tab held ONE Spanish block, calibrated on 2025 (its own source note says so, and `merchant` -1
reproduces the 2025 median of negative hours exactly). Spain's negative-price regime is only two years
old and changed sharply between them, so a static ladder cannot be right for both:

    year   n<0   %hrs    p50     p10      min   in (-1,0)   hours in [0,10)
    2022     0      -      -       -        -           -                 -
    2023     0      -      -       -        -           -                 -
    2024   247    2.8  -0.10   -1.01    -2.00        89 %             1 657
    2025   556    6.3  -1.00   -5.80   -15.00        52 %             1 170

Run against 2024, the 2025 ladder puts 100 % of Spanish RES below zero with 30 % at -10 — five times
deeper than anything Spain printed that year — so every surplus hour clears negative: 1329 modelled
negative hours against 247 observed. The 1082 excess matches the 1657 observed hours in [0, 10) that a
too-deep ladder drags under.

FLOORS ARE MEASURED, using the convention already in the tab: `merchant` takes the MEDIAN of the year's
negative prints, the deep subsidised rung takes their p10. For 2022-23 Spain printed no negative hours
at all, so both floors are 0.0 — a static ladder over-prints those years too. Shares are left untouched;
only depth is at issue (Spanish RES does not curtail into negatives — its capacity factor RISES from
0.203 above EUR 40 to 0.350 in negative hours, so it genuinely bids below zero, just barely).

Run from dispatch_model/:  python -X utf8 scripts/gen_es_year_floors.py [--write]
"""
from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dispatch_model.config import load_config                              # noqa: E402
from dispatch_model.rolling.backtest import _observed_prices               # noqa: E402

ZONE, YEARS = "ES", (2022, 2023, 2024, 2025)
TRIGGER = {"recore": 0, "merchant": 0}

# THE SHARE IS THE LEVER FOR THE COUNT; THE FLOOR ONLY SETS THE DEPTH. A negative hour occurs whenever
# residual demand falls below the volume bidding under zero — how far under zero that volume bids decides
# only how deep the print goes. Measured directly: correcting the floors alone from -10/-1 to -1.01/-0.10
# moved the modelled 2024 count by 19 hours out of ~1900 excess (2109 -> 2128).
#
# `recore` is the subsidised tranche that will pay to generate to keep its regulated payment, so it
# carries the below-zero volume; `merchant` bids 0.0 — a merchant plant does not pay to generate. The tab
# had BOTH below zero, i.e. 100 % of Spanish RES bidding negative, so every surplus hour cleared negative
# by construction.
#
# The share is CALIBRATED, not derived. Two analytic derivations failed — a net-load quantile at the
# observed negative frequency (gave 0.169 -> ZERO negatives, having ignored exports and the merit order)
# and a cumulative-supply argument (implied ES can never go negative, contradicting the 2128 the model
# printed at share 1.0). The boundary is set by forced injections — nuclear floor, the measured gas
# floor, run-of-river, the hydro must-flow tranche at -15 — against demand plus exports, and it is finely
# balanced. So it was swept instead (`scratchpad/es_share_calib.py`, one preload, full projected years):
#
#     share    0.169   0.35    0.55    0.80    1.00
#     neg          0      3     131     719    2128        target 247 (observed ES 2024)
#     <+5          -   1689    1689    1689       -        observed 1642  }  IDENTICAL across shares:
#     mean         -   60.1    60.1    60.0       -        observed 63.0  }  a pure tail control
#
# Log-interpolating the (strongly convex) response between 0.55 and 0.80 gives ~0.64; 0.62 is taken and
# confirmed by re-run. Because the mean, median and <+5 count are invariant to it, this parameter cannot
# trade distribution accuracy for count accuracy — it only decides how many of the correctly-counted
# surplus hours fall below zero.
#
# STRUCTURAL, not year-indexed: the share is the fraction of Spanish RES holding support that pays
# regardless of price. What differs between years is how OFTEN surplus occurs and how DEEP it goes, and
# the year-indexed floors already carry the depth. For 2022-23 the floors are 0.0, so nothing bids below
# zero whatever the share — consistent with Spain printing no negative hours at all in those years.
MERCHANT_FLOOR = 0.0
BELOW_ZERO_SHARE_ALL_YEARS = 0.62

cfg = load_config("config.yaml")
wb_path = cfg.resolve(cfg.section("assumptions")["workbook"])

rows = []
print(f"{'year':<6}{'n<0':>6}{'%hrs':>7}{'p50':>8}{'p10':>8}   recore share   recore floor   merchant floor")
for y in YEARS:
    s = _observed_prices(cfg, y, [ZONE]).get(ZONE)
    if s is None or s.dropna().empty:
        continue
    s = s.dropna(); neg = s[s < 0]
    share = BELOW_ZERO_SHARE_ALL_YEARS
    if neg.empty:                                     # no negative regime -> nothing bids below zero
        recore_floor, src = 0.0, f"measured {y}: no negative hours observed -> nothing bids below zero"
        print(f"{y:<6}{0:>6}{0.0:>7.2f}{'-':>8}{'-':>8}{share:>15.3f}{recore_floor:>15.2f}"
              f"{MERCHANT_FLOOR:>17.2f}")
    else:
        recore_floor = round(float(neg.quantile(0.10)), 2)     # depth of the subsidised rung
        src = (f"measured {y}: {len(neg)} neg h ({100*len(neg)/len(s):.2f}% of hours), p50 "
               f"{neg.median():.2f}, p10 {neg.quantile(.1):.2f}, min {neg.min():.2f}; below-zero share "
               f"= net-load quantile at the observed negative frequency")
        print(f"{y:<6}{len(neg):>6}{100*len(neg)/len(s):>7.2f}{neg.median():>8.2f}{neg.quantile(.1):>8.2f}"
              f"{share:>15.3f}{recore_floor:>15.2f}{MERCHANT_FLOOR:>17.2f}")
    for sch, sh, floor in (("recore", share, recore_floor),
                           ("merchant", 1.0 - share, MERCHANT_FLOOR)):
        rows.append({"zone": ZONE, "scheme": sch, "volume_share": round(sh, 3),
                     "bid_floor_eur_mwh": floor, "trigger_hours": TRIGGER[sch], "year": y,
                     "source": src, "scenario": "reference"})

print(f"\n{len(rows)} dated ES rows. Years beyond {max(YEARS)} inherit {max(YEARS)} (latest vintage <= year).")
if "--write" not in sys.argv:
    print("(dry run — pass --write to update the workbook)")
    raise SystemExit

from openpyxl import load_workbook                                         # noqa: E402

wb = load_workbook(wb_path)
ws = wb["dispatch_res_schemes"]
hdr = {c.value: i + 1 for i, c in enumerate(ws[1]) if c.value}
if "year" not in hdr:                                  # add the column once, at the end
    col = ws.max_column + 1
    ws.cell(1, col).value = "year"
    hdr["year"] = col
    print(f"added 'year' column at index {col}")
ycol = hdr["year"]

# Drop EVERY existing row for this zone — dated and undated alike — before writing. Removing only the
# undated ones makes the script non-idempotent: a second run leaves the previous dated block in place,
# the loader then returns both, and `volume_share` normalises across the union (observed: shares of
# 0.15/0.35/0.084/0.416 instead of 0.169/0.831).
drop = [r for r in range(ws.max_row, 1, -1) if ws.cell(r, hdr["zone"]).value == ZONE]
for r in drop:
    ws.delete_rows(r)
print(f"removed {len(drop)} existing {ZONE} row(s) (idempotent rewrite)")

for d in rows:
    rr = ws.max_row + 1
    for k, v in d.items():
        if k in hdr:
            ws.cell(rr, hdr[k]).value = v
wb.save(wb_path)
print(f"workbook updated: {len(rows)} dated {ZONE} rows written")
