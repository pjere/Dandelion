"""Fill NL's missing trajectory, rescale IT_NORTH wind to its bidding zone, and de-stale CH solar.

Three workbook decisions taken by the owner after the 2024 backcast and the sourcing pass. Each anchor
below carries its derivation; the underlying citations are in TYNDP_SOURCES.md.

Run from dispatch_model/:  python -X utf8 scripts/gen_tyndp_nl_it_ch.py [--write]
"""
from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dispatch_model.config import load_config                              # noqa: E402

# --------------------------------------------------------------------------------------------------
# NL — no TYNDP rows existed at all, so the projection silently used a flat +4.5 %/yr CAGR. Measured
# cost on the 2024 backcast: all 458 negative-price hours lost, +16.4 EUR/MWh, p5 at 75 vs ~0 observed.
#
# cap_solar_gw. KEV 2025 (PBL) Tabel 24 gives solar GENERATION, not capacity: 95.4 PJ = 26.5 TWh in
#   2030. Converted at a generic Dutch specific yield. That yield is not assumed — it is CONFIRMED by
#   our own lake: NL's measured 2025 capacity is 29.28 GW, and 26.5 TWh / 29.28 GW = 905 kWh/kWp. So
#   KEV's 2030 basispad is NL solar essentially FLAT from 2025, which matches its own message about grid
#   congestion and the SDE++ wind-down. 2040/2050 are EXTRAPOLATIONS (growth resumes as electrification
#   lifts demand); they are the least-supported numbers in this file.
# cap_wind_gw. "Use the projection figures" — so KEV's projected 10 GW offshore by 2030 rather than the
#   NPE's 21 GW-by-2031 target, plus ~7 GW onshore (measured 2025 total is 11.92 GW, of which ~5 GW is
#   offshore). For 2040/2050 no projection exists and only the NPE cabinet targets do (50 / 70 GW
#   offshore), so those years MIX a target basis into a projection series — flagged deliberately.
# demand_twh. KEV 2025 Tabel 24 "Elektriciteitsbalans", row Totaal verbruik, at 1 PJ = 0.277778 TWh:
#   2024 = 425 PJ = 118.1 TWh, 2030 = 548 PJ = 152.2 TWh. 2050 = 273 TWh is the NPE's indicative DIRECT
#   electricity demand — a different accounting basis from KEV's total consumption, so 2040 is an
#   interpolation across two bases, i.e. an educated guess as instructed.
# cap_flex_gw. LEFT ABSENT by instruction: NL therefore gets no #83 adequacy block.
# --------------------------------------------------------------------------------------------------
NL = {
    "cap_solar_gw": {2025: 29.28, 2030: 29.4, 2040: 45.0, 2050: 60.0},
    "cap_wind_gw":  {2025: 11.92, 2030: 17.0, 2040: 58.0, 2050: 78.0},
    "demand_twh":   {2025: 122.0, 2030: 152.2, 2040: 205.0, 2050: 273.0},
}

# --------------------------------------------------------------------------------------------------
# IT_NORTH cap_wind_gw — confirmed by the workbook owner as a NATIONAL figure entered against a bidding
# zone. Measured Nord capacity at end-2019 is 0.129 GW (GSE FER 2019 §3.3.6: 10'715 MW national x 1.2 %
# Nord share), so the entered 3.0 GW at 2025 implied 23x growth in six years in a zone with poor wind
# resource and no pipeline — Italian wind is in the South (Puglia 24.0 %, Sicilia 17.7 %).
#
# The entered series is kept as a GROWTH SHAPE (x1 / x2 / x3.33 / x4.67 relative to its first anchor)
# and re-based on the measured Nord capacity, rather than simply multiplied by the 1.2 % share — that
# would give 0.036 GW in 2025, BELOW the 2019 measurement, i.e. a decline. Re-basing preserves the
# scenario's growth intent at a scale the zone can actually carry.
# --------------------------------------------------------------------------------------------------
IT_WIND_2025 = 0.15          # measured 0.129 GW in 2019, near-flat since — modest growth to 2025
IT_NORTH = {"cap_wind_gw": {2025: IT_WIND_2025, 2030: IT_WIND_2025 * 2.0,
                            2040: round(IT_WIND_2025 * 3.33, 3), 2050: round(IT_WIND_2025 * 4.67, 3)}}

# --------------------------------------------------------------------------------------------------
# CH cap_solar_gw — the 5.0 GW 2025 anchor was already passed on the BFE's own series before 2024
# (measured 2019 = 2.498 GW; BFE reports 21'458 TJ = 5'961 GWh of PV production in 2024, i.e. ~7 GW of
# fleet at Swiss yields). Near-term anchors corrected with a DECAYING CAGR; the long-term intent of the
# scenario (25 GW by 2040, 35 GW by 2050) is deliberately left untouched, since only the near term was
# stale. Resulting CAGRs: 12 %/yr 2025-30, 5.2 %/yr 2030-40, 3.4 %/yr 2040-50.
# --------------------------------------------------------------------------------------------------
CH = {"cap_solar_gw": {2025: 8.5, 2030: 15.0}}          # 2040 / 2050 unchanged at 25 / 35

PLAN = {"NL": NL, "IT_NORTH": IT_NORTH, "CH": CH}

cfg = load_config("config.yaml")
wb_path = cfg.resolve(cfg.section("assumptions")["workbook"])
sheet = pd.read_excel(wb_path, sheet_name="dispatch_tyndp")

rows = []
for zone, plan in PLAN.items():
    for var, series in plan.items():
        cur = sheet[(sheet["zone"] == zone) & (sheet["variable"] == var)]
        have = {int(r.year): float(r.value) for r in cur.itertuples()}
        print(f"\n{zone}/{var}")
        for y, v in sorted(series.items()):
            old = have.get(y)
            print(f"    {y}  {('—' if old is None else f'{old:g}'):>8} -> {v:g}"
                  + ("   (new)" if old is None else "   (replaced)"))
            rows.append({"zone": zone, "variable": var, "year": y, "value": v})

print(f"\n{len(rows)} anchors to write")
print("NOTE: run gen_tyndp_baseline.py afterwards — NL now has future anchors, so its 2019 rows will be")
print("      filled automatically from measured ENTSO-E data (solar 7.23 GW, wind 4.48 GW, 113.9 TWh).")

if "--write" not in sys.argv:
    print("\n(dry run — pass --write to update the workbook)")
    raise SystemExit

from openpyxl import load_workbook                                         # noqa: E402

wb = load_workbook(wb_path)
ws = wb["dispatch_tyndp"]
hdr = {c.value: i + 1 for i, c in enumerate(ws[1])}
zc, vc, yc, valc = hdr["zone"], hdr["variable"], hdr["year"], hdr["value"]
index = {(ws.cell(r, zc).value, ws.cell(r, vc).value, int(ws.cell(r, yc).value)): r
         for r in range(2, ws.max_row + 1) if ws.cell(r, yc).value is not None}
n_e = n_a = 0
for r_ in rows:
    key = (r_["zone"], r_["variable"], r_["year"])
    if key in index:
        ws.cell(index[key], valc).value = r_["value"]; n_e += 1
    else:
        rr = ws.max_row + 1
        ws.cell(rr, zc).value = r_["zone"]; ws.cell(rr, vc).value = r_["variable"]
        ws.cell(rr, yc).value = r_["year"]; ws.cell(rr, valc).value = r_["value"]
        index[key] = rr; n_a += 1
wb.save(wb_path)
print(f"workbook updated: {n_e} replaced, {n_a} added")
