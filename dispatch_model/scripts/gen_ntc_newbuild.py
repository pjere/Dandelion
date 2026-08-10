"""Build the `dispatch_ntc_newbuild` tab: interconnector commissioning steps, per project.

WHY THIS SHAPE. An interconnector is a discrete asset, exactly like a reactor in `dispatch_nuclear_newbuild`
— it commissions on a date, adds a fixed MW, and is flat either side. It is NOT a quantity that grows
smoothly, so it must not be modelled as `dispatch_tyndp` models capacity: a ratio linearly interpolated
between anchors would invent capacity in every year nothing was built, and would spread one cable's rating
across a decade. Hence one row per PROJECT with a commissioning year, and a STEP lookup downstream.

WHY A DELTA AND NOT A LEVEL. The projection clears a target year from a reference year's hourly NTC series
(`assemble.hourly_ntc`, which carries real outages and flow-based domain shrinkage). That series already
embodies the grid as it stood in the reference year, so what a projection needs is only what was built
SINCE: delta = Σ(projects with ref_year < commissioning_year ≤ target_year), added to the reference series.
This is why no baseline capacity table is needed — and that matters, because ENTSO-E publishes none. ACER's
opinion on the 2024 draft TYNDP explicitly *asks* ENTSO-E to start publishing the current and starting grid
per boundary and direction, which is as clear a statement as one gets that it is not available today. Any
ratio formulation would have required inventing that baseline; a delta formulation does not.

Capacity is applied SYMMETRICALLY to both directions. New interconnectors are overwhelmingly HVDC links
with one rating, and the asymmetry in the `NTC` table comes from network constraints around the border
rather than from the wire.

THREE CERTAINTY TIERS, in the `scenario` column:

  built             already commissioned since the 2019 reference year. Every one is VALIDATED in-house
                    against the step in observed max hourly flow (`entsoe_flows`) — see the evidence
                    column on each row below. These are not forecasts; omitting them is a known error,
                    because the projection's reference year is 2019 and these wires exist today.
  reference         committed / under construction, individually sourced.
  tyndp_candidate   ENTSO-E TYNDP 2024 assessed investment candidates. NOT committed — these are the
                    projects the CBA exists to decide on. Written to the tab but OFF by default
                    (`tyndp.load_ntc_newbuild` takes built+reference), so enabling them is a deliberate
                    high-build scenario rather than an accident.

Run from dispatch_model/:  python -X utf8 scripts/gen_ntc_newbuild.py [--write]
"""
from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dispatch_model.config import load_config                              # noqa: E402
from dispatch_model.rolling.assemble import NTC                            # noqa: E402

SHEET = "dispatch_ntc_newbuild"

# --- tier 1: built since the 2019 reference year -----------------------------------------------------
# `evidence` is the step in max hourly flow across the border, measured on `entsoe_flows` (MW, fwd/bwd).
# entsoe_flows holds no 2020/2021 rows for these borders, so a commissioning year between two measured
# years is bracketed rather than pinned; the MW step itself is unambiguous in every case.
BUILT = [
    ("DE_LU", "BE", "ALEGrO", 1000, 2020,
     "2019 249/312 -> 2022 1164/1298; first DE-BE link, border was AC-coupled only before"),
    ("FR", "GB", "IFA2", 1000, 2021, "FR-GB 2019 2036 -> 2022 3071"),
    ("FR", "GB", "ElecLink", 1000, 2022, "FR-GB 2022 3071 -> 2023 4087"),
    ("FR", "IT_NORTH", "Savoie-Piemont", 1200, 2023,
     "FR->IT 2019 3563 -> 2024 4554; half capacity 2022-11, full 2023-08 (RTE)"),
    ("DK", "GB", "Viking Link", 1400, 2023, "DK-GB 2022 0/0 -> 2023 1408/1454"),
]

# --- tier 2: committed / under construction ----------------------------------------------------------
REFERENCE = [
    ("FR", "ES", "Bay of Biscay (Golfe de Gascogne)", 2200, 2028,
     "INELFE (RTE/REE) + EIB EUR 1.6 bn 2025; doubles FR-ES to 5000 MW. Corroborated by TYNDP 2024 "
     "reference grid ES00-FR00 = 5000/5000, i.e. today's 2800 + 2200"),
]

# --- tier 3: ENTSO-E TYNDP 2024 assessed investment candidates ---------------------------------------
# Source: "20231103 - Electricity and Hydrogen Reference Grid & Investment Candidates.xlsx", sheet
# "3. Elec Invest Candidates", column "DIRECT CAPACITY INCREASE (MW)", summed per border and YEAR over the
# 80 rows whose FROM/TO nodes both map into a modelled zone. Scenario column is "All" on every row.
# These sit ON TOP of the TYNDP 2030 reference grid, which already contains the committed projects above.
TYNDP_CANDIDATES = [
    ("CH", "AT_SI", 1000, 2030), ("DE_LU", "AT_SI", 2000, 2030), ("DE_LU", "AT_SI", 2000, 2040),
    ("IT_NORTH", "AT_SI", 1000, 2030), ("IT_NORTH", "AT_SI", 500, 2035),
    ("IT_NORTH", "AT_SI", 1150, 2040),
    ("DE_LU", "BE", 1000, 2030), ("DE_LU", "BE", 3000, 2040),
    ("FR", "BE", 1000, 2030), ("FR", "BE", 2000, 2040),
    ("BE", "GB", 1400, 2035), ("BE", "GB", 2000, 2040),
    ("BE", "NL", 2000, 2030), ("BE", "NL", 2000, 2035),
    ("DE_LU", "CH", 1000, 2030), ("DE_LU", "CH", 300, 2035), ("DE_LU", "CH", 2000, 2040),
    ("FR", "CH", 2000, 2040),
    ("CH", "IT_NORTH", 2000, 2030), ("CH", "IT_NORTH", 1000, 2035),
    ("DE_LU", "DK", 1000, 2030), ("DE_LU", "DK", 4000, 2040),
    ("FR", "DE_LU", 2000, 2040),
    ("DE_LU", "NL", 1000, 2030), ("DE_LU", "NL", 1000, 2035), ("DE_LU", "NL", 1000, 2040),
    ("DE_LU", "PL_CZ", 7000, 2040),
    ("DK", "GB", 1400, 2035),
    ("FR", "ES", 3000, 2030), ("FR", "ES", 3000, 2040),
    ("ES", "PT", 2000, 2040),
    ("FR", "GB", 1250, 2035), ("FR", "GB", 2600, 2040),
    ("FR", "IT_NORTH", 2000, 2030),
    ("NL", "GB", 2000, 2030),
    ("IT_NORTH", "IT_SOUTH", 4000, 2030),
]

#: TYNDP candidates on boundaries the model has no border for — reported, never silently dropped.
UNMAPPED = [("AT_SI", "PL_CZ", 1000, 2030), ("AT_SI", "PL_CZ", 500, 2040),
            ("DE_LU", "GB", 1400, 2030), ("DK", "NL", 2000, 2040)]


def _key(a: str, b: str) -> str:
    """Border label in the same orientation `assemble.NTC` uses, so the two never disagree."""
    if (a, b) in NTC:
        return f"{a}-{b}"
    if (b, a) in NTC:
        return f"{b}-{a}"
    return f"{a}-{b}"                                     # unmapped; reported separately


def build_rows() -> list[dict]:
    rows = []
    for a, b, proj, mw, yr, ev in BUILT:
        rows.append({"border": _key(a, b), "project": proj, "capacity_mw": mw,
                     "commissioning_year": yr, "scenario": "built", "source": ev})
    for a, b, proj, mw, yr, src in REFERENCE:
        rows.append({"border": _key(a, b), "project": proj, "capacity_mw": mw,
                     "commissioning_year": yr, "scenario": "reference", "source": src})
    for a, b, mw, yr in TYNDP_CANDIDATES:
        rows.append({"border": _key(a, b), "project": f"TYNDP 2024 candidate {yr}", "capacity_mw": mw,
                     "commissioning_year": yr, "scenario": "tyndp_candidate",
                     "source": "ENTSO-E TYNDP 2024, sheet '3. Elec Invest Candidates'"})
    return rows


if __name__ == "__main__":
    cfg = load_config("config.yaml")
    wb_path = cfg.resolve(cfg.section("assumptions")["workbook"])
    rows = build_rows()
    df = pd.DataFrame(rows)

    print(f"{len(df)} rows for `{SHEET}`\n")
    for sc in ("built", "reference", "tyndp_candidate"):
        d = df[df.scenario == sc]
        print(f"--- {sc}: {len(d)} rows, {d.capacity_mw.sum():.0f} MW total ---")
        for _, r in d.iterrows():
            print(f"  {r.border:<18}{r.project:<34}{r.capacity_mw:>6.0f} MW  {r.commissioning_year}")
        print()

    bad = sorted({r["border"] for r in rows} - {f"{a}-{b}" for a, b in NTC})
    print(f"borders not in the model's NTC table: {bad or 'none'}")
    print("TYNDP candidates dropped for want of a model border:")
    for a, b, mw, yr in UNMAPPED:
        print(f"  {a}-{b:<10}{mw:>6} MW  {yr}   (model carries no such border)")

    if "--write" not in sys.argv:
        print("\n(dry run — pass --write to update the workbook)")
        raise SystemExit

    from openpyxl import load_workbook                                     # noqa: E402
    wbk = load_workbook(wb_path)
    if SHEET in wbk.sheetnames:                    # rebuild wholesale: this tab is generated, not edited
        del wbk[SHEET]
    ws = wbk.create_sheet(SHEET)
    cols = ["border", "project", "capacity_mw", "commissioning_year", "scenario", "source"]
    ws.append(cols)
    for r in rows:
        ws.append([r[c] for c in cols])
    wbk.save(wb_path)
    print(f"\nworkbook updated: `{SHEET}` written with {len(rows)} rows")
