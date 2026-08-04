"""Restate `cap_hydro_gw` on ONE definition across all zones, and split pumped storage into `cap_psp_gw`.

THE DEFINITION. `cap_hydro_gw` means RESERVOIR + RUN-OF-RIVER, EXCLUDING pumped storage. `tyndp._CAP_VAR`
already encodes it (`hydro_psp` maps to its own `cap_psp_gw`), and the dispatch forces it: PSP is not in
the hydro merit order at all — `flexibility.storage` sizes the storage LP from the zone's measured
`hydro_psp` stack capacity and dispatches it as storage. A PSP-inclusive factor scales a stack that does
not contain the PSP it is counting.

THE SHEET DID NOT FOLLOW IT UNIFORMLY. Measured against 2019 ENTSO-E installed capacity, FR (26 GW) and
CH (15 GW) were entered PSP-inclusive; DE_LU (5 GW) and ES (20 GW) on the correct res+ror basis. One tech
set therefore could not baseline every zone — PSP-inclusive gave DE 0.34, PSP-exclusive gave FR 1.41 —
which is why `cap_hydro_gw` sat excluded from `gen_tyndp_baseline.py` and permanently clamped.

WHAT THIS DOES. For the two deviating zones, subtract measured pumped-storage capacity from every
`cap_hydro_gw` anchor and write that PSP into `cap_psp_gw`, so all zones end up on one definition and the
factor means the same thing everywhere.

PSP LEVEL USED, per the workbook owner's decision:
  * FR — static. France has no new pumped storage; measured PSP is flat at ~5.05 GW across 2019-2026.
  * CH — the 2022 level, held constant for the future. Swiss PSP is NOT static: Nant de Drance added
    ~0.9 GW in 2022, so the 2019 figure would understate the fleet for every projected year. The 2019
    row keeps its own measured value, so the PSP factor itself carries that step.

Provenance and the full argument are in TYNDP_SOURCES.md.

Run from dispatch_model/:  python -X utf8 scripts/fix_tyndp_hydro_basis.py [--write]
"""
from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dispatch_model.config import load_config                              # noqa: E402
from dispatch_model.io.entsoe_hist import load_installed_capacity          # noqa: E402
from dispatch_model.neighbours.blocks import constituents                  # noqa: E402

REF_YEAR = 2019
#: zone -> year whose measured PSP capacity represents the fleet for the PROJECTED years.
#: FR is flat so the choice is immaterial; CH is not (Nant de Drance, 2022).
PSP_FUTURE_YEAR = {"FR": 2022, "CH": 2022}
#: zones whose `cap_hydro_gw` anchors are PSP-INCLUSIVE and must be restated
RESTATE = ("FR", "CH")


def psp_gw(zone: str, year: int) -> float:
    tot = 0.0
    for z in constituents(zone):
        cap = load_installed_capacity(cfg, z, year) or {}
        tot += float(cap.get("hydro_psp", 0.0))
    return round(tot / 1e3, 3)


cfg = load_config("config.yaml")
wb_path = cfg.resolve(cfg.section("assumptions")["workbook"])
sheet = pd.read_excel(wb_path, sheet_name="dispatch_tyndp")

edits, adds = [], []
for zone in RESTATE:
    g = sheet[(sheet["zone"] == zone) & (sheet["variable"] == "cap_hydro_gw")].sort_values("year")
    if g.empty:
        print(f"{zone}: no cap_hydro_gw rows — nothing to restate")
        continue
    p_ref = psp_gw(zone, REF_YEAR)
    p_fut = psp_gw(zone, PSP_FUTURE_YEAR[zone])
    print(f"\n{zone}: PSP measured {REF_YEAR} = {p_ref} GW, {PSP_FUTURE_YEAR[zone]} = {p_fut} GW "
          f"(used for all projected years)")
    print(f"  {'year':>6}{'was (incl PSP)':>16}{'PSP':>8}{'becomes (res+ror)':>20}")
    for r in g.itertuples():
        y, was = int(r.year), float(r.value)
        p = p_ref if y <= REF_YEAR else p_fut
        now = round(was - p, 3)
        if now <= 0:
            print(f"  {y:>6}{was:>16g}{p:>8g}{'NEGATIVE — skipped':>20}")
            continue
        print(f"  {y:>6}{was:>16g}{p:>8g}{now:>20g}")
        edits.append({"zone": zone, "variable": "cap_hydro_gw", "year": y, "value": now})
        adds.append({"zone": zone, "variable": "cap_psp_gw", "year": y, "value": p})

if not edits:
    print("\nnothing to do")
    raise SystemExit

print(f"\n{len(edits)} cap_hydro_gw values restated, {len(adds)} cap_psp_gw rows to write")
if "--write" not in sys.argv:
    print("(dry run — pass --write to update the workbook)")
    raise SystemExit

from openpyxl import load_workbook                                         # noqa: E402

wb = load_workbook(wb_path)
ws = wb["dispatch_tyndp"]
hdr = {c.value: i + 1 for i, c in enumerate(ws[1])}
zc, vc, yc, valc = hdr["zone"], hdr["variable"], hdr["year"], hdr["value"]
index = {(ws.cell(r, zc).value, ws.cell(r, vc).value, int(ws.cell(r, yc).value)): r
         for r in range(2, ws.max_row + 1) if ws.cell(r, yc).value is not None}
n_e = n_a = 0
for r_ in edits + adds:
    key = (r_["zone"], r_["variable"], r_["year"])
    if key in index:
        ws.cell(index[key], valc).value = r_["value"]
        n_e += 1
    else:
        rr = ws.max_row + 1
        ws.cell(rr, zc).value = r_["zone"]; ws.cell(rr, vc).value = r_["variable"]
        ws.cell(rr, yc).value = r_["year"]; ws.cell(rr, valc).value = r_["value"]
        index[key] = rr
        n_a += 1
wb.save(wb_path)
print(f"workbook updated: {n_e} replaced, {n_a} added — rationale in TYNDP_SOURCES.md")
