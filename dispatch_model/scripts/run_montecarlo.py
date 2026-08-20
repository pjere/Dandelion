"""Monte-Carlo projection: N independent weather draws, each carried coherently through steps iii-vii.

For every draw d this runs, in order:

  1. weathergen simulates a COMPLETE new 20-year hourly trajectory (42 stations, 8 variables), seeded by
     SeedSequence([master_seed, d]) so draw d is reproducible regardless of order or of which draws ran;
  2. `POWERSIM_WEATHER_CUBE` points every downstream model at that cube — demand (iii), RES (iv),
     availability (v) and the dispatch preload all read the same file, which is what makes the draw
     coherent rather than four unrelated randomisations;
  3. step v re-runs on that weather, so plant availability (thermal derating, reservoir wetness) follows
     the draw instead of being frozen;
  4. the 20-year dispatch runs;
  5. deliverables are written to `<out>/draw_<ddd>/`, with the draw's own weather summary as a tab.

**Parallel across draws.** Each draw is fully isolated: its own weather cube, its own step-v availability
(lake layer `availability_mc`, partition `wdraw=d`, selected by `DISPATCH_AVAIL_WDRAW`), its own demand/RES
driver cache (lake partition `realization=d`) and its own output directory. Nothing is shared and written,
so `--workers K` runs K draws concurrently. Pick K against RAM, not cores: each worker holds a full
projection preload plus its HiGHS model, and the machine this was built on fits about 3.

**Resumable.** A draw whose analysis workbook already exists is skipped, so an interrupted ensemble
continues where it stopped. That matters: one draw is ~4.5 h with the flexibility module on.

    python -u -X utf8 -W ignore scripts/run_montecarlo.py --draws 50 --workers 3 [--out reports/mc]
"""
from __future__ import annotations

import argparse
import os
import subprocess
import sys
import time
from pathlib import Path

_HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(_HERE.parent))
sys.path.insert(0, str(_HERE))

_ROOT = _HERE.parents[1]


def _run(cmd: list[str], env: dict, label: str) -> None:
    t0 = time.monotonic()
    r = subprocess.run(cmd, env=env, cwd=str(_HERE.parent))
    if r.returncode != 0:
        raise RuntimeError(f"{label} failed with exit code {r.returncode}")
    print(f"    [{label}] {time.monotonic() - t0:.0f}s", flush=True)


def run_draw(draw: int, out_root: Path, master_seed: int, keep_cube: bool,
             n_weeks: int | None = None) -> None:
    from mc_weather import generate_cube, weather_summary

    ddir = out_root / f"draw_{draw:03d}"
    if (ddir / "projection_20y_analysis.xlsx").exists():
        print(f"[mc] draw {draw}: already complete, skipping", flush=True)
        return
    print(f"[mc] draw {draw}: starting", flush=True)
    t0 = time.monotonic()

    cube = Path("scratchpad/mc") / f"cube_{draw:03d}.nc"
    if not cube.exists():
        tw = time.monotonic()
        generate_cube(draw, cube, master_seed)
        print(f"    [weathergen] {time.monotonic() - tw:.0f}s -> {cube}", flush=True)

    env = dict(os.environ)
    env["POWERSIM_WEATHER_CUBE"] = str((_HERE.parent / cube).resolve())
    env["DISPATCH_CAPTURE_DISPATCH"] = "1"
    env["DISPATCH_AVAIL_WDRAW"] = str(draw)      # this draw's availability, not the shared tables

    # step v on THIS draw's weather: one availability realization per weather realization
    _run([sys.executable, "-u", "-X", "utf8", "-W", "ignore", "-c",
          "import sys; sys.path.insert(0, r'%s');"
          "from availability_model.config import load_config;"
          "from availability_model.projection.engine import project;"
          "project(load_config(r'%s'), n_draws=1, layer='availability_mc', "
          "partitions={'wdraw': %d}, draw_ids=[%d])"
          % (str(_ROOT / "availability_model"), str(_ROOT / "availability_model" / "config.yaml"),
             draw, draw)],
         env, "step-v availability")

    work = Path("scratchpad/mc") / f"proj_{draw:03d}"
    # `--draw d` is what keys the demand/RES driver caches (lake partition `realization=d`). Passing 0 for
    # every draw would make them share ONE cache entry, and the mtime freshness test would then hand draw
    # d+1 the drivers computed for draw d whenever its cube predates that cache — a silent wrong answer.
    # The index no longer selects weather (the cube itself differs per draw), so it is free to use as a key.
    cmd = [sys.executable, "-u", "-X", "utf8", "-W", "ignore",
           "scripts/run_projection_20y.py", "--out", str(work), "--draw", str(draw)]
    if n_weeks:
        cmd += ["--n-weeks", str(n_weeks)]
    _run(cmd, env, "dispatch 20y")

    _run([sys.executable, "-u", "-X", "utf8", "-W", "ignore",
          "scripts/build_projection_deliverables.py", "--src", str(work), "--out", str(ddir)],
         env, "deliverables")

    # the draw's own weather, as a tab — without it a price path cannot be attributed to a winter
    import pandas as pd
    from openpyxl import load_workbook
    ws = weather_summary(_HERE.parent / cube)
    ws.to_csv(ddir / "weather_summary.csv", index=False)
    xl = ddir / "projection_20y_analysis.xlsx"
    with pd.ExcelWriter(xl, engine="openpyxl", mode="a", if_sheet_exists="replace") as w:
        ws.round(2).to_excel(w, sheet_name="Weather (this draw)", index=False)
    wb = load_workbook(xl)
    if "Weather (this draw)" in wb.sheetnames:                      # place it just after the Notice
        wb.move_sheet("Weather (this draw)", offset=-(len(wb.sheetnames) - 2))
        wb.save(xl)

    if not keep_cube:
        cube.unlink(missing_ok=True)                                # 472 MB each; 50 of them is 23 GB
    print(f"[mc] draw {draw}: done in {(time.monotonic() - t0) / 60:.1f} min -> {ddir}", flush=True)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--draws", type=int, default=50, help="number of draws (0..N-1)")
    ap.add_argument("--start-draw", type=int, default=0)
    ap.add_argument("--out", default="reports/mc")
    ap.add_argument("--master-seed", type=int, default=20260629)
    ap.add_argument("--keep-cubes", action="store_true", help="retain each 472 MB cube after the draw")
    ap.add_argument("--workers", type=int, default=1,
                    help="draws in flight. Bound by RAM (a preload + HiGHS model each), not by cores")
    ap.add_argument("--n-weeks", type=int, default=None,
                    help="truncate each projected year — smoke-testing the chain, not a real ensemble")
    a = ap.parse_args()

    out_root = Path(a.out)
    out_root.mkdir(parents=True, exist_ok=True)
    todo = list(range(a.start_draw, a.draws))
    print(f"[mc] {len(todo)} draw(s) {todo[0]}..{todo[-1]}, master_seed={a.master_seed}, out={out_root}",
          flush=True)
    t0 = time.monotonic()
    if a.workers <= 1:
        for d in todo:
            run_draw(d, out_root, a.master_seed, a.keep_cubes, a.n_weeks)
    else:
        from concurrent.futures import ProcessPoolExecutor, as_completed
        done, failed = 0, []
        with ProcessPoolExecutor(max_workers=a.workers) as ex:
            futs = {ex.submit(run_draw, d, out_root, a.master_seed, a.keep_cubes, a.n_weeks): d
                    for d in todo}
            for f in as_completed(futs):
                d = futs[f]
                try:
                    f.result()
                except Exception as exc:                                   # noqa: BLE001
                    # one bad draw must not lose the ensemble: record it, keep the rest, report at the end
                    failed.append((d, f"{type(exc).__name__}: {exc}"))
                    print(f"[mc] draw {d} FAILED: {type(exc).__name__}: {exc}", flush=True)
                done += 1
                print(f"[mc] {done}/{len(todo)} draws finished "
                      f"({(time.monotonic() - t0) / 3600:.2f} h elapsed)", flush=True)
        if failed:
            print(f"[mc] {len(failed)} draw(s) failed: {failed}", flush=True)
            print("[mc] re-run the same command to retry them (completed draws are skipped)", flush=True)
    print(f"[mc] ensemble complete in {(time.monotonic() - t0) / 3600:.2f} h", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
