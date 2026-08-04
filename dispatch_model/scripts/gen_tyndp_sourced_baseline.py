"""Write the 2019 anchors that CANNOT be measured from the lake, sourced from official statistics.

Companion to `gen_tyndp_baseline.py`, which derives 2019 anchors from ENTSO-E installed capacity. Three
zone/variable pairs are beyond its reach and it reports them as "no actual available":

  CH/cap_solar_gw, CH/cap_wind_gw   ENTSO-E's Swiss submission carries only hydro and nuclear — there is
                                    no solar or wind row in ANY year. The generation proxy is invalid
                                    too: p99.9 of 2019 CH solar generation is 297 MW against a ~2.5 GW
                                    fleet, because Swiss PV is behind the meter.
  IT_NORTH/cap_solar_gw, cap_wind_gw  entsoe_installed_capacity has ZERO rows for IT_NORTH, and ENTSO-E
                                    publishes capacity per country, not per bidding zone.

Without a 2019 row `_interp(2019)` clamps flat to the first anchor (2025), the factor collapses to ~1.0,
and RES is FROZEN at its reference-year level for every projected year — indistinguishable in the output
from a deliberate flat scenario. Measured cost on the 2024 backcast: CH lost all 292 negative-price hours
and ran +20.3 €/MWh; IT_NORTH ran with a frozen solar fleet.

Every figure here is quoted from a named table in an official publication. Full provenance, including
the arithmetic behind the two sums, is in `TYNDP_SOURCES.md` — read it before changing any number.

Run from dispatch_model/:  python -X utf8 scripts/gen_tyndp_sourced_baseline.py [--write]
"""
from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dispatch_model.config import load_config                              # noqa: E402

REF_YEAR = 2019

#: (zone, variable) -> (GW at end-2019, one-line provenance). See TYNDP_SOURCES.md for the full citation.
SOURCED = {
    ("CH", "cap_solar_gw"): (
        2.498,
        "BFE/Swissolar, Markterhebung Sonnenenergie 2019, §4.3 'Gesamthaft installierte Leistung in kW "
        "per Ende Jahr', row Photovoltaik Total, col 2019 = 2'498'050 kWp"),
    ("CH", "cap_wind_gw"): (
        0.075,
        "~75 MW at end-2019 (~40 turbines). WEAKEST figure here: the BFE wind page gives no MW total and "
        "the renewables statistics report wind in TJ, not MW. 2.9% of CH wind+solar, so a +/-20% error "
        "moves the res factor <0.6%. Replace from BFE renewables statistics Anhang B when available."),
    ("IT_NORTH", "cap_solar_gw"): (
        9.2625,
        "GSE, Solare Fotovoltaico - Rapporto Statistico 2019, p.21, sum of the Nord bidding-zone regions: "
        "Piemonte 1642.5 + Valle d'Aosta 24.6 + Lombardia 2398.8 + Trentino-AA 442.7 + Veneto 1995.8 + "
        "Friuli VG 545.2 + Liguria 112.8 + Emilia-Romagna 2100.1 = 9262.5 MW"),
    ("IT_NORTH", "cap_wind_gw"): (
        0.129,
        "GSE, Rapporto Statistico FER 2019, §3.3.6 p.59: national 10'715 MW x Nord regional shares "
        "(Piemonte 0.2 + Liguria 0.5 + Emilia-Romagna 0.4 + Veneto 0.1 = 1.2%) = 128.6 MW. Lombardia, "
        "Valle d'Aosta, Trentino-AA and Friuli VG have no installed wind."),
}

cfg = load_config("config.yaml")
wb_path = cfg.resolve(cfg.section("assumptions")["workbook"])
sheet = pd.read_excel(wb_path, sheet_name="dispatch_tyndp")

print(f"{'zone':<10}{'variable':<16}{'2019':>8}{'first anchor':>16}{'implied ratio':>15}   status")
print("-" * 100)
rows, warn = [], []
for (zone, var), (val, prov) in SOURCED.items():
    g = sheet[(sheet["zone"] == zone) & (sheet["variable"] == var)]
    future = g[g["year"] > REF_YEAR].sort_values("year")
    if future.empty:
        print(f"{zone:<10}{var:<16}{val:>8.3f}{'—':>16}{'—':>15}   SKIP: no future anchor to scale to")
        continue
    y0, v0 = int(future.iloc[0].year), float(future.iloc[0].value)
    existing = g[g["year"] == REF_YEAR]
    ratio = v0 / val if val > 0 else float("inf")
    status = "replaces existing" if not existing.empty else "new row"
    print(f"{zone:<10}{var:<16}{val:>8.3f}{f'{y0}={v0:g}':>16}{ratio:>14.2f}x   {status}")
    rows.append({"zone": zone, "variable": var, "year": REF_YEAR, "value": val})
    # An implausible implied growth usually means the FUTURE anchor is wrong, not the baseline — the
    # baseline is measured, the anchor is a scenario input. Say so rather than silently writing it.
    if ratio > 10:
        warn.append(f"{zone}/{var}: {y0} anchor {v0:g} GW implies {ratio:.0f}x growth from the measured "
                    f"{REF_YEAR} value of {val:g} GW — check whether that anchor is a NATIONAL figure "
                    f"entered against a bidding zone")

print("\nprovenance:")
for (zone, var), (val, prov) in SOURCED.items():
    print(f"  {zone}/{var} = {val} GW\n      {prov}")

if warn:
    print("\n⚠ implausible implied growth — the SCENARIO anchor is the suspect, not the measured baseline:")
    for w in warn:
        print(f"  {w}")

if "--write" not in sys.argv:
    print("\n(dry run — pass --write to update the workbook)")
    raise SystemExit

from openpyxl import load_workbook                                         # noqa: E402

wb = load_workbook(wb_path)
ws = wb["dispatch_tyndp"]
hdr = {c.value: i + 1 for i, c in enumerate(ws[1])}
zc, vc, yc, valc = hdr["zone"], hdr["variable"], hdr["year"], hdr["value"]
existing = {(ws.cell(r, zc).value, ws.cell(r, vc).value, int(ws.cell(r, yc).value)): r
            for r in range(2, ws.max_row + 1) if ws.cell(r, yc).value is not None}
repl = add = 0
for r_ in rows:
    key = (r_["zone"], r_["variable"], REF_YEAR)
    if key in existing:
        ws.cell(existing[key], valc).value = r_["value"]
        repl += 1
    else:
        rr = ws.max_row + 1
        ws.cell(rr, zc).value = r_["zone"]
        ws.cell(rr, vc).value = r_["variable"]
        ws.cell(rr, yc).value = REF_YEAR
        ws.cell(rr, valc).value = r_["value"]
        add += 1
wb.save(wb_path)
print(f"\nworkbook updated: {repl} replaced, {add} added — provenance recorded in TYNDP_SOURCES.md")
